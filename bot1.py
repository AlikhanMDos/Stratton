from openpyxl import Workbook
from datetime import datetime
import random
import os


# Excel file
wb = Workbook()
ws = wb.active
ws.title = "TDSheet"

# Columns of the sheet
ws['A1'] = "Имя"
ws['B1'] = "Текущая дата"
ws['C1'] = "Текущее время"

# Random name generator
names = ["Alikhan", "Yernur", "Alinur", "Maksat", "Yersultan", "Andrey", "Aisulu", "Rasul", "Alex", "Tom"]

for _ in range(10):
    generate_random_name = random.choice(names)
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    ws.append([generate_random_name, current_date, current_time])

random_number = random.randint(100, 999)
filename = f"{generate_random_name}_{current_date}_{random_number}.xlsx"



# Path to save
folder_path = os.path.join(os.path.expanduser("~"), "Documents", "skcu")
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

path_to_save = os.path.join(folder_path, filename)
wb.save(path_to_save)

print("File successfully created:", path_to_save)
